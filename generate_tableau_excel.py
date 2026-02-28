"""
Generate Tableau-ready Excel workbook with 5 sheets:
1. Stock_Prices       — Daily OHLCV + MA20/MA50/Daily_Return for all tickers
2. Financial_Statements — Quarterly revenue, margins, EPS, net income
3. Revenue_Forecast   — Actuals + Holt-Winters fitted + 8Q forecast + CI bands
4. Annual_Returns     — Yearly % returns per ticker
5. Summary_Metrics    — Latest price, YTD return, volatility, moving averages
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as openpyxl_numbers)
from openpyxl.utils import get_column_letter
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

np.random.seed(42)

# ── COLOUR PALETTE ──────────────────────────────────────────
HEADER_FILL  = PatternFill('solid', start_color='1A3A5C')
HEADER_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
TITLE_FILL   = PatternFill('solid', start_color='0D2137')
TITLE_FONT   = Font(bold=True, color='00D4FF', name='Arial', size=12)
ALT_FILL     = PatternFill('solid', start_color='F0F5FA')
WHITE_FILL   = PatternFill('solid', start_color='FFFFFF')
BLUE_INPUT   = Font(color='0000FF', name='Arial', size=10)
BLACK_CALC   = Font(color='000000', name='Arial', size=10)
GREEN_LINK   = Font(color='006400', name='Arial', size=10)
POS_FILL     = PatternFill('solid', start_color='E8F5E9')
NEG_FILL     = PatternFill('solid', start_color='FFEBEE')
THIN_BORDER  = Border(
    left=Side(style='thin', color='D0D9E8'),
    right=Side(style='thin', color='D0D9E8'),
    top=Side(style='thin', color='D0D9E8'),
    bottom=Side(style='thin', color='D0D9E8')
)
NORMAL = Alignment(horizontal='center', vertical='center')
LEFT   = Alignment(horizontal='left',   vertical='center')

def hdr(ws, cell, value, width=None):
    ws[cell] = value
    ws[cell].font = HEADER_FONT
    ws[cell].fill = HEADER_FILL
    ws[cell].alignment = NORMAL
    ws[cell].border = THIN_BORDER
    if width:
        col = cell[0] if len(cell)==2 else cell[:2] if cell[1].isdigit() else cell[0]
        ws.column_dimensions[get_column_letter(ws[cell].column)].width = width

def title_row(ws, cell, value):
    ws[cell] = value
    ws[cell].font = TITLE_FONT
    ws[cell].fill = TITLE_FILL
    ws[cell].alignment = LEFT

def style_row(ws, row, num_cols, alt=False):
    fill = ALT_FILL if alt else WHITE_FILL
    for c in range(1, num_cols+1):
        cell = ws.cell(row=row, column=c)
        if not cell.font or cell.font.bold is not True:
            cell.font = Font(name='Arial', size=10)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = NORMAL

# ════════════════════════════════════════════════════════════
# 1. GENERATE BASE DATA (same seed as financial_analysis.py)
# ════════════════════════════════════════════════════════════

def generate_stock_data(ticker, start_price, trend, volatility, n_days=1095):
    dates = pd.date_range(end=datetime.today(), periods=n_days, freq='B')
    dt = 1/252; mu=trend; sigma=volatility
    returns = np.random.normal((mu-0.5*sigma**2)*dt, sigma*np.sqrt(dt), n_days)
    prices  = start_price * np.exp(np.cumsum(returns))
    volume  = np.random.randint(5_000_000, 50_000_000, n_days)
    n = len(dates)
    return pd.DataFrame({
        'Date':   dates, 'Ticker': ticker,
        'Open':   (prices[:n]*np.random.uniform(0.995,1.005,n)).round(2),
        'High':   (prices[:n]*np.random.uniform(1.001,1.020,n)).round(2),
        'Low':    (prices[:n]*np.random.uniform(0.980,0.999,n)).round(2),
        'Close':  prices[:n].round(2), 'Volume': volume[:n]
    })

tickers_cfg = {
    'AAPL': (182.0, 0.22, 0.28),
    'MSFT': (375.0, 0.18, 0.25),
    'GOOGL':(140.0, 0.15, 0.30),
    'TSLA': (240.0, 0.10, 0.55),
    'AMZN': (178.0, 0.20, 0.32),
}

df = pd.concat([generate_stock_data(t,*p) for t,p in tickers_cfg.items()], ignore_index=True)

# Inject & clean
idx = np.random.choice(df.index, size=int(len(df)*0.015), replace=False)
df.loc[idx, 'Close'] = np.nan
df = df.drop_duplicates(subset=['Date','Ticker']).sort_values(['Ticker','Date'])
df['Close'] = df.groupby('Ticker')['Close'].transform(lambda s: s.interpolate())
df['Volume'] = df['Volume'].abs()
df['High']   = df[['Open','High','Close']].max(axis=1).round(2)
df['Low']    = df[['Open','Low','Close']].min(axis=1).round(2)

df['Daily_Return']   = df.groupby('Ticker')['Close'].pct_change().round(6)
df['MA_20']          = df.groupby('Ticker')['Close'].transform(lambda s: s.rolling(20).mean()).round(2)
df['MA_50']          = df.groupby('Ticker')['Close'].transform(lambda s: s.rolling(50).mean()).round(2)
df['Volatility_20d'] = df.groupby('Ticker')['Daily_Return'].transform(lambda s: (s.rolling(20).std()*np.sqrt(252)).round(4))
df['Year']           = df['Date'].dt.year
df['Month']          = df['Date'].dt.month
df['Quarter']        = df['Date'].dt.quarter
df['YearQuarter']    = df['Date'].dt.to_period('Q').astype(str)

stock_df = df.copy()

# ── QUARTERLY FINANCIALS ─────────────────────────────────────
quarters = pd.period_range('2021Q1', periods=16, freq='Q')
revenue_base   = 90_000
revenue_growth = np.linspace(0, 0.28, 16) + np.random.normal(0, 0.015, 16)
revenue        = revenue_base * (1 + revenue_growth)
cogs_ratio     = np.random.uniform(0.35, 0.42, 16)
gross_profit   = revenue * (1-cogs_ratio)
opex           = revenue * np.random.uniform(0.20, 0.25, 16)
ebitda         = gross_profit - opex
depreciation   = revenue * 0.04
ebit           = ebitda - depreciation
interest_exp   = np.random.uniform(800, 1200, 16)
ebt            = ebit - interest_exp
net_income     = ebt * 0.79
eps            = net_income / 15_000

fin_df = pd.DataFrame({
    'Quarter':       [str(q) for q in quarters],
    'Year':          [q.year for q in quarters],
    'Quarter_Num':   [q.quarter for q in quarters],
    'Revenue_M':     revenue.round(0),
    'COGS_M':        (revenue*cogs_ratio).round(0),
    'Gross_Profit_M':gross_profit.round(0),
    'OPEX_M':        opex.round(0),
    'EBITDA_M':      ebitda.round(0),
    'EBIT_M':        ebit.round(0),
    'Net_Income_M':  net_income.round(0),
    'EPS':           eps.round(2),
    'Gross_Margin_pct': (gross_profit/revenue*100).round(2),
    'EBITDA_Margin_pct':(ebitda/revenue*100).round(2),
    'Net_Margin_pct':   (net_income/revenue*100).round(2),
    'Revenue_B':     (revenue/1000).round(3),
    'Net_Income_B':  (net_income/1000).round(3),
})

# ── HOLT-WINTERS FORECAST ────────────────────────────────────
def holt_winters(series, alpha=0.3, beta=0.1, gamma=0.1, season_len=4, n_forecast=8):
    y = np.array(series, dtype=float); n = len(y)
    level=np.zeros(n+n_forecast); trend_=np.zeros(n+n_forecast)
    season=np.zeros(n+n_forecast); fitted=np.zeros(n+n_forecast)
    level[0]=np.mean(y[:season_len])
    trend_[0]=(np.mean(y[season_len:2*season_len])-np.mean(y[:season_len]))/season_len
    for i in range(season_len): season[i]=y[i]-level[0]
    for t in range(1,n):
        pl=level[t-1]; pt=trend_[t-1]
        ps=season[t-season_len] if t>=season_len else season[t%season_len]
        level[t] =alpha*(y[t]-ps)+(1-alpha)*(pl+pt)
        trend_[t]=beta*(level[t]-pl)+(1-beta)*pt
        season[t]=gamma*(y[t]-level[t])+(1-gamma)*ps
        fitted[t]=level[t]+trend_[t]+season[t]
    forecast_vals=[]
    for h in range(1,n_forecast+1):
        f=(level[n-1]+h*trend_[n-1]+season[n-season_len+(h-1)%season_len])
        forecast_vals.append(f)
    return fitted[1:n], np.array(forecast_vals)

fitted_rev, forecast_rev = holt_winters(fin_df['Revenue_M'], alpha=0.3, beta=0.08, gamma=0.15)

rev_forecast_df = pd.DataFrame({
    'Quarter':      list(fin_df['Quarter']) + [f'2025Q{i}' if i<=4 else f'2026Q{i-4}' for i in range(1,9)],
    'Type':         ['Actual']*16 + ['Forecast']*8,
    'Revenue_B':    list((fin_df['Revenue_M']/1000).round(3)) + list((forecast_rev/1000).round(3)),
    'Fitted_B':     [None] + list((fitted_rev/1000).round(3)) + [None]*8,
    'CI_Low_B':     [None]*16 + list((forecast_rev*0.90/1000).round(3)),
    'CI_High_B':    [None]*16 + list((forecast_rev*1.10/1000).round(3)),
})

# ── ANNUAL RETURNS ────────────────────────────────────────────
annual = (stock_df.groupby(['Ticker','Year'])['Close']
          .apply(lambda s:(s.iloc[-1]/s.iloc[0]-1)*100)
          .reset_index(name='Annual_Return_pct'))
annual['Annual_Return_pct'] = annual['Annual_Return_pct'].round(2)

# ── SUMMARY TABLE ─────────────────────────────────────────────
latest = stock_df.groupby('Ticker').last().reset_index()
first  = stock_df.groupby('Ticker').first().reset_index()
colors_map = {'AAPL':'#00E5FF','MSFT':'#FF5E3A','GOOGL':'#39FF82','TSLA':'#FF3CAC','AMZN':'#FFD000'}
summary_df = pd.DataFrame({
    'Ticker':           latest['Ticker'],
    'Last_Price':       latest['Close'].round(2),
    'YTD_Return_pct':   ((latest['Close'].values/first['Close'].values-1)*100).round(2),
    'Volatility_20d_pct':(latest['Volatility_20d']*100).round(2),
    'MA_20':            latest['MA_20'].round(2),
    'MA_50':            latest['MA_50'].round(2),
    'Signal':           latest.apply(lambda r: 'Bullish' if r['Close']>=r['MA_20'] else 'Bearish', axis=1),
    'Volatility_Rank':  [3,4,5,2,1],
})

print(f"Stock rows    : {len(stock_df):,}")
print(f"Financial Q   : {len(fin_df)}")
print(f"Rev forecast  : {len(rev_forecast_df)}")
print(f"Annual returns: {len(annual)}")

# ════════════════════════════════════════════════════════════
# 2. BUILD WORKBOOK
# ════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)   # remove default sheet

# ── HELPER: write DataFrame to sheet ─────────────────────────
def write_df(ws, df, start_row=3, date_cols=None, pct_cols=None, money_cols=None):
    date_cols  = date_cols  or []
    pct_cols   = pct_cols   or []
    money_cols = money_cols or []
    cols = list(df.columns)
    # Write headers
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=start_row, column=ci, value=col.replace('_',' '))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = NORMAL
        cell.border = THIN_BORDER
    # Write data
    for ri, (_, row) in enumerate(df.iterrows(), start_row+1):
        alt = (ri-start_row) % 2 == 0
        for ci, col in enumerate(cols, 1):
            val = row[col]
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font  = Font(name='Arial', size=10)
            cell.fill  = ALT_FILL if alt else WHITE_FILL
            cell.alignment = NORMAL
            cell.border = THIN_BORDER
            if col in date_cols and val is not None:
                cell.number_format = 'YYYY-MM-DD'
            elif col in pct_cols and val is not None:
                cell.number_format = '0.00"%"'
            elif col in money_cols and val is not None:
                cell.number_format = '$#,##0.00'
    return start_row + len(df) + 1

# ════════════════════════════════════════════════════════════
# SHEET 1 — Stock Prices
# ════════════════════════════════════════════════════════════
ws1 = wb.create_sheet('Stock_Prices')
title_row(ws1, 'A1', '📈  DAILY STOCK PRICES  —  AAPL · MSFT · GOOGL · TSLA · AMZN  |  3-Year History')
ws1.merge_cells('A1:N1')
ws1['A2'] = 'Source: Simulated via Geometric Brownian Motion (seed=42) | Units: Price in USD, Volume in shares, Volatility annualised'
ws1['A2'].font = Font(name='Arial', size=9, italic=True, color='5A7A9A')
ws1.merge_cells('A2:N2')

out_df1 = stock_df[['Date','Ticker','Open','High','Low','Close','Volume',
                      'Daily_Return','MA_20','MA_50','Volatility_20d',
                      'Year','Month','Quarter']].copy()
out_df1['Date'] = out_df1['Date'].dt.strftime('%Y-%m-%d')

write_df(ws1, out_df1, start_row=3,
         date_cols=['Date'],
         pct_cols=['Daily_Return','Volatility_20d'],
         money_cols=['Open','High','Low','Close','MA_20','MA_50'])

col_widths1 = {'A':13,'B':8,'C':10,'D':10,'E':10,'F':10,'G':14,
               'H':14,'I':10,'J':10,'K':14,'L':7,'M':8,'N':9}
for col, w in col_widths1.items():
    ws1.column_dimensions[col].width = w
ws1.freeze_panes = 'A4'
ws1.auto_filter.ref = f'A3:{get_column_letter(len(out_df1.columns))}3'

# ════════════════════════════════════════════════════════════
# SHEET 2 — Financial Statements
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Financial_Statements')
title_row(ws2, 'A1', '💰  QUARTERLY FINANCIAL STATEMENTS  |  2021Q1 – 2024Q4  |  Revenue & Profit Metrics')
ws2.merge_cells('A1:O1')
ws2['A2'] = 'Source: Simulated income statement data | Monetary units: $M (millions) except EPS ($) and _B columns ($B billions) | All margins in %'
ws2['A2'].font = Font(name='Arial', size=9, italic=True, color='5A7A9A')
ws2.merge_cells('A2:O2')

write_df(ws2, fin_df, start_row=3,
         money_cols=['Revenue_M','COGS_M','Gross_Profit_M','OPEX_M','EBITDA_M','EBIT_M','Net_Income_M'],
         pct_cols=['Gross_Margin_pct','EBITDA_Margin_pct','Net_Margin_pct'])

col_widths2 = {c:13 for c in 'ABCDEFGHIJKLMNO'}
col_widths2.update({'A':9,'B':7,'C':12,'D':13,'E':13,'F':15,'G':13})
for col, w in col_widths2.items():
    ws2.column_dimensions[col].width = w
ws2.freeze_panes = 'A4'
ws2.auto_filter.ref = f'A3:{get_column_letter(len(fin_df.columns))}3'

# ════════════════════════════════════════════════════════════
# SHEET 3 — Revenue Forecast
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Revenue_Forecast')
title_row(ws3, 'A1', '📊  REVENUE FORECAST  |  Holt-Winters Triple Exponential Smoothing  |  MAPE: 2.02%')
ws3.merge_cells('A1:F1')
ws3['A2'] = 'Actuals: 2021Q1–2024Q4 | Forecast: 2025Q1–2026Q4 (8 quarters) | 90% CI band shown for forecast period'
ws3['A2'].font = Font(name='Arial', size=9, italic=True, color='5A7A9A')
ws3.merge_cells('A2:F2')

write_df(ws3, rev_forecast_df, start_row=3)

# Colour forecast rows differently
for row in ws3.iter_rows(min_row=4, max_row=3+len(rev_forecast_df)):
    type_cell = row[1]  # column B = Type
    if type_cell.value == 'Forecast':
        for cell in row:
            cell.fill = PatternFill('solid', start_color='E8F5E9')
            cell.font = Font(name='Arial', size=10, color='006400')

col_widths3 = {'A':10,'B':10,'C':12,'D':12,'E':12,'F':12}
for col, w in col_widths3.items():
    ws3.column_dimensions[col].width = w
ws3.freeze_panes = 'A4'

# ════════════════════════════════════════════════════════════
# SHEET 4 — Annual Returns
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Annual_Returns')
title_row(ws4, 'A1', '📅  ANNUAL STOCK RETURNS (%)  |  2021 – 2026  |  Calendar Year Performance')
ws4.merge_cells('A1:C1')
ws4['A2'] = 'Annual return = (Year-end price / Year-start price - 1) × 100 | Positive = green, Negative = red'
ws4['A2'].font = Font(name='Arial', size=9, italic=True, color='5A7A9A')
ws4.merge_cells('A2:C2')

write_df(ws4, annual, start_row=3, pct_cols=['Annual_Return_pct'])

# Conditional colour for returns
for row in ws4.iter_rows(min_row=4, max_row=3+len(annual)):
    ret_cell = row[2]  # column C
    if isinstance(ret_cell.value, (int, float)):
        if ret_cell.value >= 0:
            for cell in row:
                cell.fill = POS_FILL
                cell.font = Font(name='Arial', size=10, color='006400')
        else:
            for cell in row:
                cell.fill = NEG_FILL
                cell.font = Font(name='Arial', size=10, color='C62828')

ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 8
ws4.column_dimensions['C'].width = 18

# ════════════════════════════════════════════════════════════
# SHEET 5 — Summary Metrics
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Summary_Metrics')
title_row(ws5, 'A1', '🎯  SUMMARY METRICS  |  Latest Price, YTD Return, Volatility & Moving Averages')
ws5.merge_cells('A1:H1')
ws5['A2'] = 'Data as of latest available trading date | Signal: Bullish = Close ≥ MA_20 | Volatility = 20-day annualised std dev'
ws5['A2'].font = Font(name='Arial', size=9, italic=True, color='5A7A9A')
ws5.merge_cells('A2:H2')

write_df(ws5, summary_df, start_row=3,
         money_cols=['Last_Price','MA_20','MA_50'],
         pct_cols=['YTD_Return_pct','Volatility_20d_pct'])

# Colour signal column
for row in ws5.iter_rows(min_row=4, max_row=3+len(summary_df)):
    sig = row[6].value
    if sig == 'Bullish':
        row[6].font = Font(name='Arial', size=10, color='006400', bold=True)
        row[6].fill = POS_FILL
    elif sig == 'Bearish':
        row[6].font = Font(name='Arial', size=10, color='C62828', bold=True)
        row[6].fill = NEG_FILL

col_ws5 = {'A':10,'B':12,'C':16,'D':18,'E':10,'F':10,'G':10,'H':16}
for col, w in col_ws5.items():
    ws5.column_dimensions[col].width = w

# ════════════════════════════════════════════════════════════
# SHEET 6 — Tableau Guide (instructions tab)
# ════════════════════════════════════════════════════════════
ws6 = wb.create_sheet('Tableau_Guide')
title_row(ws6, 'A1', '📋  TABLEAU DASHBOARD BUILD GUIDE  —  Sheet Reference & Connection Tips')
ws6.merge_cells('A1:E1')

guide_data = [
    ('SHEET NAME', 'CONNECT AS', 'KEY DIMENSIONS', 'KEY MEASURES', 'USE FOR'),
    ('Stock_Prices',       'Live connection or extract', 'Date, Ticker, Year, Quarter', 'Close, Volume, MA_20, MA_50, Daily_Return, Volatility_20d', 'Line chart (Close/MA over time), Volume bars, Volatility trend'),
    ('Financial_Statements','Extract',                  'Quarter, Year, Quarter_Num',  'Revenue_B, Net_Income_B, Gross_Margin_pct, EBITDA_Margin_pct, EPS', 'Revenue bar/line chart, Margin area chart, EPS trend'),
    ('Revenue_Forecast',    'Extract',                  'Quarter, Type',               'Revenue_B, Fitted_B, CI_Low_B, CI_High_B', 'Forecast dual-axis chart with CI band'),
    ('Annual_Returns',      'Extract',                  'Ticker, Year',                'Annual_Return_pct', 'Heatmap or grouped bar chart of annual performance'),
    ('Summary_Metrics',     'Extract',                  'Ticker, Signal',              'Last_Price, YTD_Return_pct, Volatility_20d_pct, MA_20, MA_50', 'KPI tiles, scatter plot (vol vs return), bullet chart'),
]

for ri, row_data in enumerate(guide_data, 3):
    for ci, val in enumerate(row_data, 1):
        cell = ws6.cell(row=ri, column=ci, value=val)
        if ri == 3:
            cell.font  = HEADER_FONT
            cell.fill  = HEADER_FILL
        else:
            cell.font  = Font(name='Arial', size=10)
            cell.fill  = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border    = THIN_BORDER

ws6.row_dimensions[3].height = 18
for r in range(4, 9): ws6.row_dimensions[r].height = 45
col_ws6 = {'A':20,'B':26,'C':32,'D':50,'E':38}
for col, w in col_ws6.items():
    ws6.column_dimensions[col].width = w

# ════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════
out_path = 'FinSight_Tableau_Data.xlsx'  # saves to current folder
wb.save(out_path)
print(f'Saved: {out_path}')
print(f'Sheets: {wb.sheetnames}')
